"""
报表渲染引擎

支持数据透视表和自由布局两种模式，
可导出为图片、Excel、PDF。
"""

import io
from typing import Dict, Any, List, Optional
from dataclasses import dataclass, field

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages


# 聚合函数映射
AGG_FUNCTIONS = {
    "sum": np.sum,
    "mean": np.mean,
    "count": len,
    "max": np.max,
    "min": np.min,
    "std": np.std,
    "median": np.median,
}

AGG_NAMES = {
    "sum": "求和",
    "mean": "平均",
    "count": "计数",
    "max": "最大",
    "min": "最小",
    "std": "标准差",
    "median": "中位数",
}


@dataclass
class CellConfig:
    """单元格配置（自由布局模式）"""
    row: int = 0
    col: int = 0
    row_span: int = 1
    col_span: int = 1
    type: str = "text"  # text, field, header, title
    value: str = ""  # 文本内容或字段名
    bound_field: Optional[str] = None  # 绑定的数据字段
    agg: str = "sum"  # 聚合方式
    format: str = ""  # 数字格式
    style: Dict[str, Any] = field(default_factory=dict)


@dataclass
class PivotConfig:
    """数据透视表配置"""
    row_fields: List[str] = field(default_factory=list)
    col_fields: List[str] = field(default_factory=list)
    value_fields: List[str] = field(default_factory=list)
    aggregations: Dict[str, str] = field(default_factory=dict)  # field -> agg
    filters: Dict[str, Any] = field(default_factory=dict)  # field -> value


@dataclass
class ChartConfig:
    """图表配置"""
    chart_type: str = "bar"  # bar, pie, line, area, scatter, horizontal_bar, donut
    x_field: str = ""  # 分类/X轴字段
    y_field: str = ""  # 数值/Y轴字段
    group_field: str = ""  # 分组字段（可选，多系列）
    agg: str = "sum"  # Y轴聚合方式
    title: str = "图表"
    color_scheme: str = "default"  # default, pastel, dark, bright
    show_values: bool = True  # 是否显示数值标签
    top_n: int = 0  # 只显示前N项，0表示全部
    filters: List[Dict[str, Any]] = field(default_factory=list)  # 从探索视图继承的筛选契约
    time_unit: str = "none"  # none, day, week, month, quarter, year
    bins: int = 20  # 连续分组轴的分箱数
    discovery_note: str = ""  # 探索阶段保留的解释边界


@dataclass
class ReportConfig:
    """报表配置"""
    mode: str = "pivot"  # "pivot" 或 "free"
    title: str = "报表"
    pivot: PivotConfig = field(default_factory=PivotConfig)
    cells: List[CellConfig] = field(default_factory=list)
    row_count: int = 12
    col_count: int = 8
    charts: List[ChartConfig] = field(default_factory=list)
    styles: Dict[str, Any] = field(default_factory=lambda: {
        "header_bg": "#2E86AB",
        "header_color": "#FFFFFF",
        "grid_color": "#E0E4E8",
        "font_size": 11,
        "title_font_size": 16,
    })


class PivotBuilder:
    """数据透视表构建器"""

    @staticmethod
    def build(df: pd.DataFrame, config: PivotConfig) -> pd.DataFrame:
        """
        构建数据透视表

        Returns:
            渲染后的 DataFrame（交叉表形式）
        """
        if df.empty:
            return pd.DataFrame()

        # 过滤数据
        filtered_df = df.copy()
        for field, value in config.filters.items():
            if field in filtered_df.columns:
                filtered_df = filtered_df[filtered_df[field] == value]

        # 如果没有值字段，返回空表
        if not config.value_fields:
            return pd.DataFrame({"提示": ["请至少添加一个值字段"]})

        # 构建 aggfunc 字典
        aggfunc = {}
        for field in config.value_fields:
            agg = config.aggregations.get(field, "sum")
            func = AGG_FUNCTIONS.get(agg, np.sum)
            aggfunc[field] = func

        # 使用 pd.pivot_table
        try:
            if config.row_fields and config.col_fields:
                pivot = pd.pivot_table(
                    filtered_df,
                    values=config.value_fields,
                    index=config.row_fields,
                    columns=config.col_fields,
                    aggfunc=aggfunc,
                    fill_value=0,
                    margins=True,
                    margins_name="总计",
                )
            elif config.row_fields:
                pivot = pd.pivot_table(
                    filtered_df,
                    values=config.value_fields,
                    index=config.row_fields,
                    aggfunc=aggfunc,
                    fill_value=0,
                )
            elif config.col_fields:
                pivot = pd.pivot_table(
                    filtered_df,
                    values=config.value_fields,
                    columns=config.col_fields,
                    aggfunc=aggfunc,
                    fill_value=0,
                )
            else:
                # 只有值字段，做整体聚合
                result = {}
                for field in config.value_fields:
                    agg = config.aggregations.get(field, "sum")
                    func = AGG_FUNCTIONS.get(agg, np.sum)
                    result[field] = [func(filtered_df[field])]
                pivot = pd.DataFrame(result, index=["总计"])

            # 格式化列名
            if isinstance(pivot.columns, pd.MultiIndex):
                pivot.columns = [
                    " | ".join(str(c) for c in col if pd.notna(c))
                    for col in pivot.columns.values
                ]

            # 格式化索引名
            if isinstance(pivot.index, pd.MultiIndex):
                pivot.index = [
                    " | ".join(str(i) for i in idx if pd.notna(i))
                    for idx in pivot.index.values
                ]

            return pivot.reset_index()

        except Exception as e:
            return pd.DataFrame({"错误": [f"透视表构建失败: {str(e)}"]})


class ChartBuilder:
    """图表构建器"""

    COLOR_SCHEMES = {
        "default": ["#2E86AB", "#A23B72", "#F18F01", "#C73E1D", "#28A745", "#6C757D", "#17a2b8", "#ffc107"],
        "pastel": ["#FFB3BA", "#FFDFBA", "#FFFFBA", "#BAFFC9", "#BAE1FF", "#E6E6FA", "#F0E68C", "#DDA0DD"],
        "dark": ["#1f4e79", "#7030a0", "#c55a11", "#548235", "#44546a", "#7f6000", "#5b9bd5", "#70ad47"],
        "bright": ["#FF6B6B", "#4ECDC4", "#45B7D1", "#FFA07A", "#98D8C8", "#F7DC6F", "#BB8FCE", "#85C1E9"],
    }

    @staticmethod
    def prepare_frame(df: pd.DataFrame, config: ChartConfig) -> pd.DataFrame:
        """Reapply the exploration grain contract before static rendering."""
        frame = df.copy()
        for item in config.filters or []:
            field_name = str(item.get("field", ""))
            if field_name not in frame.columns:
                continue
            if item.get("kind") == "range":
                numeric = pd.to_numeric(frame[field_name], errors="coerce")
                lower = item.get("min")
                upper = item.get("max")
                mask = pd.Series(True, index=frame.index)
                if lower is not None:
                    mask &= numeric >= float(lower)
                if upper is not None:
                    mask &= numeric <= float(upper)
                frame = frame.loc[mask]
            elif item.get("kind") == "in":
                allowed = {str(value) for value in item.get("values", [])}
                frame = frame.loc[frame[field_name].astype("string").isin(allowed)]

        if config.x_field in frame.columns and config.agg != "none":
            x_values = frame[config.x_field]
            if pd.api.types.is_numeric_dtype(x_values) and x_values.nunique(dropna=True) > max(2, int(config.bins)):
                frame[config.x_field] = pd.cut(
                    pd.to_numeric(x_values, errors="coerce"),
                    bins=max(2, int(config.bins)),
                    duplicates="drop",
                ).astype("string")
            elif config.time_unit != "none":
                dates = pd.to_datetime(x_values, errors="coerce")
                period = {
                    "day": "D", "week": "W", "month": "M",
                    "quarter": "Q", "year": "Y",
                }.get(config.time_unit)
                if period:
                    frame[config.x_field] = dates.dt.to_period(period).dt.start_time
        return frame

    @staticmethod
    def aggregate_frame(frame: pd.DataFrame, config: ChartConfig) -> pd.DataFrame:
        """Aggregate including a first-class record-count measure."""
        group_fields = [config.x_field]
        if config.group_field and config.group_field in frame.columns:
            group_fields.append(config.group_field)
        if config.y_field == "__count__":
            grouped = frame.groupby(group_fields, observed=True, dropna=False).size().reset_index(name="__count__")
        else:
            agg_func = config.agg if config.agg in AGG_FUNCTIONS else "sum"
            grouped = frame.groupby(group_fields, observed=True, dropna=False)[config.y_field].agg(agg_func).reset_index()
        if len(group_fields) > 1:
            return grouped.pivot(index=config.x_field, columns=config.group_field, values=config.y_field).fillna(0)
        return grouped.set_index(config.x_field)

    @staticmethod
    def build(df: pd.DataFrame, config: ChartConfig) -> plt.Figure:
        """
        根据配置渲染图表
        Returns: matplotlib Figure
        """
        if df.empty or not config.x_field or not config.y_field:
            fig, ax = plt.subplots(figsize=(8, 5))
            ax.text(0.5, 0.5, "请配置图表字段", ha="center", va="center", fontsize=14)
            ax.axis("off")
            return fig

        if config.x_field not in df.columns or (config.y_field != "__count__" and config.y_field not in df.columns):
            fig, ax = plt.subplots(figsize=(8, 5))
            ax.text(0.5, 0.5, f"字段不存在: {config.x_field} 或 {config.y_field}", ha="center", va="center", fontsize=12)
            ax.axis("off")
            return fig

        # 准备数据，并严格重放探索阶段的筛选与粒度
        colors = ChartBuilder.COLOR_SCHEMES.get(config.color_scheme, ChartBuilder.COLOR_SCHEMES["default"])
        render_df = ChartBuilder.prepare_frame(df, config)
        pivot_data = ChartBuilder.aggregate_frame(render_df, config)

        # 取 top_n
        if config.top_n > 0 and len(pivot_data) > config.top_n:
            if isinstance(pivot_data, pd.DataFrame):
                top_idx = pivot_data.sum(axis=1).nlargest(config.top_n).index
            else:
                top_idx = pivot_data.nlargest(config.top_n).index
            pivot_data = pivot_data.loc[top_idx]

        fig, ax = plt.subplots(figsize=(10, 6))

        if config.chart_type == "bar":
            if isinstance(pivot_data, pd.DataFrame) and pivot_data.shape[1] > 1:
                pivot_data.plot(kind="bar", ax=ax, color=colors, width=0.8)
            else:
                vals = pivot_data.values.flatten() if isinstance(pivot_data, pd.DataFrame) else pivot_data.values
                bars = ax.bar(range(len(pivot_data)), vals, color=colors[0], width=0.6)
                ax.set_xticks(range(len(pivot_data)))
                ax.set_xticklabels(pivot_data.index, rotation=45, ha="right")
                if config.show_values:
                    for bar in bars:
                        height = bar.get_height()
                        ax.text(bar.get_x() + bar.get_width()/2., height,
                                f"{height:.0f}", ha="center", va="bottom", fontsize=9)

        elif config.chart_type == "horizontal_bar":
            if isinstance(pivot_data, pd.DataFrame) and pivot_data.shape[1] > 1:
                pivot_data.plot(kind="barh", ax=ax, color=colors, width=0.8)
            else:
                vals = pivot_data.values.flatten() if isinstance(pivot_data, pd.DataFrame) else pivot_data.values
                bars = ax.barh(range(len(pivot_data)), vals, color=colors[0], height=0.6)
                ax.set_yticks(range(len(pivot_data)))
                ax.set_yticklabels(pivot_data.index)
                if config.show_values:
                    for bar in bars:
                        width = bar.get_width()
                        ax.text(width, bar.get_y() + bar.get_height()/2.,
                                f"{width:.0f}", ha="left", va="center", fontsize=9)

        elif config.chart_type == "pie":
            vals = pivot_data.values.flatten() if isinstance(pivot_data, pd.DataFrame) else pivot_data.values
            wedges, texts, autotexts = ax.pie(
                vals, labels=pivot_data.index, autopct="%1.1f%%",
                colors=colors, startangle=90
            )
            for autotext in autotexts:
                autotext.set_fontsize(9)

        elif config.chart_type == "donut":
            vals = pivot_data.values.flatten() if isinstance(pivot_data, pd.DataFrame) else pivot_data.values
            wedges, texts, autotexts = ax.pie(
                vals, labels=pivot_data.index, autopct="%1.1f%%",
                colors=colors, startangle=90, wedgeprops=dict(width=0.4)
            )
            for autotext in autotexts:
                autotext.set_fontsize(9)

        elif config.chart_type == "line":
            if isinstance(pivot_data, pd.DataFrame) and pivot_data.shape[1] > 1:
                for i, col in enumerate(pivot_data.columns):
                    ax.plot(pivot_data.index, pivot_data[col], marker="o", label=col, color=colors[i % len(colors)])
                ax.legend()
            else:
                vals = pivot_data.values.flatten() if isinstance(pivot_data, pd.DataFrame) else pivot_data.values
                ax.plot(pivot_data.index, vals, marker="o", color=colors[0], linewidth=2)
            ax.set_xticklabels(pivot_data.index, rotation=45, ha="right")

        elif config.chart_type == "area":
            if isinstance(pivot_data, pd.DataFrame) and pivot_data.shape[1] > 1:
                pivot_data.plot(kind="area", ax=ax, color=colors, alpha=0.6)
            else:
                vals = pivot_data.values.flatten() if isinstance(pivot_data, pd.DataFrame) else pivot_data.values
                ax.fill_between(range(len(pivot_data)), vals, alpha=0.6, color=colors[0])
                ax.plot(range(len(pivot_data)), vals, color=colors[0], linewidth=2)
                ax.set_xticks(range(len(pivot_data)))
                ax.set_xticklabels(pivot_data.index, rotation=45, ha="right")

        elif config.chart_type == "scatter":
            if config.group_field and config.group_field in render_df.columns:
                for index, (name, subset) in enumerate(render_df.groupby(config.group_field, observed=True, dropna=False)):
                    ax.scatter(subset[config.x_field], subset[config.y_field], alpha=0.6, color=colors[index % len(colors)], s=50, label=str(name))
                ax.legend()
            else:
                ax.scatter(render_df[config.x_field], render_df[config.y_field], alpha=0.6, color=colors[0], s=50)
            ax.set_xlabel(config.x_field)

        # 通用样式
        ax.set_title(config.title, fontsize=14, fontweight="bold", pad=15)
        y_label = "记录数" if config.y_field == "__count__" else config.y_field
        ax.set_ylabel(y_label if config.chart_type not in ("pie", "donut") else "")
        ax.set_xlabel(config.x_field if config.chart_type not in ("pie", "donut") else "")
        ax.grid(axis="y", alpha=0.3, linestyle="--")
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

        plt.tight_layout()
        return fig

    @staticmethod
    def build_to_bytes(df: pd.DataFrame, config: ChartConfig) -> bytes:
        """渲染图表并返回 PNG 字节"""
        fig = ChartBuilder.build(df, config)
        buf = io.BytesIO()
        fig.savefig(buf, format="png", bbox_inches="tight", dpi=150)
        buf.seek(0)
        plt.close(fig)
        return buf.getvalue()


class FreeReportBuilder:
    """自由布局报表构建器"""

    @staticmethod
    def build(df: pd.DataFrame, cells: List[CellConfig]) -> pd.DataFrame:
        """
        构建自由布局报表
        将单元格配置转换为可渲染的 DataFrame
        """
        if not cells:
            return pd.DataFrame({"提示": ["请添加单元格"]})

        max_row = max(c.row + c.row_span for c in cells)
        max_col = max(c.col + c.col_span for c in cells)

        # 创建网格
        grid = [["" for _ in range(max_col)] for _ in range(max_row)]

        for cell in cells:
            if cell.type == "field" and cell.bound_field and cell.bound_field in df.columns:
                agg = AGG_FUNCTIONS.get(cell.agg, np.sum)
                try:
                    value = agg(df[cell.bound_field].dropna())
                    if pd.isna(value):
                        value = "N/A"
                    else:
                        value = FreeReportBuilder._format_value(value, cell.format)
                except Exception:
                    # 限定 Exception 避免吞掉 KeyboardInterrupt / SystemExit
                    value = "N/A"
                grid[cell.row][cell.col] = str(value)
            elif cell.type == "title":
                grid[cell.row][cell.col] = cell.value
            elif cell.type == "header":
                grid[cell.row][cell.col] = cell.value
            else:
                grid[cell.row][cell.col] = cell.value

        return pd.DataFrame(grid)

    @staticmethod
    def _format_value(value, fmt: str) -> str:
        if fmt == "#,##0":
            return f"{int(value):,}"
        elif fmt == "#,##0.00":
            return f"{value:,.2f}"
        elif fmt == "0%":
            return f"{value:.1%}"
        elif fmt == "0.00%":
            return f"{value:.2%}"
        else:
            return str(round(value, 2)) if isinstance(value, (int, float)) else str(value)


class ReportExporter:
    """报表导出器"""

    @staticmethod
    def to_image(df: pd.DataFrame, config: ReportConfig, chart_df: Optional[pd.DataFrame] = None) -> bytes:
        """渲染为 PNG 图片字节（支持表格+图表组合）"""
        styles = config.styles
        n_rows, n_cols = df.shape
        has_table = n_rows > 0 and n_cols > 0
        has_charts = len(config.charts) > 0 and chart_df is not None

        # 大数据量分页处理：超过50行只显示前50行
        MAX_ROWS = 50
        original_rows = n_rows
        if has_table and n_rows > MAX_ROWS:
            df = df.head(MAX_ROWS)
            n_rows = MAX_ROWS

        # 高 DPI 保证清晰度
        DPI = 300

        # 根据数据量动态计算尺寸和字体
        cell_height = 0.55  # 增大单元格高度
        cell_width = 2.2    # 增大单元格宽度
        title_height = 0.8 if config.title else 0
        base_font = max(9, min(14, 18 - n_cols * 0.3))  # 列多自动缩小字体
        base_font = max(7, base_font - max(0, n_rows - 20) * 0.05)  # 行多也适当缩小

        # 计算布局
        if has_charts and has_table:
            n_charts = len(config.charts)
            table_h = max(4, n_rows * cell_height + title_height + 1)
            chart_h = n_charts * 5
            fig_height = min(40, table_h + chart_h + 2)
            fig_width = min(24, max(12, n_cols * cell_width + 2))
        elif has_charts:
            n_charts = len(config.charts)
            cols_per_row = min(2, n_charts)
            rows_of_charts = (n_charts + cols_per_row - 1) // cols_per_row
            fig_height = min(30, max(6, rows_of_charts * 5 + 1))
            fig_width = min(20, max(10, cols_per_row * 5))
        else:
            # 纯表格：按数据量动态计算，保证每行有足够高度
            fig_height = min(32, max(5, n_rows * cell_height + title_height + 1.5))
            fig_width = min(24, max(10, n_cols * cell_width + 2))

        fig = plt.figure(figsize=(fig_width, fig_height))

        if has_charts and has_table:
            table_height_ratio = min(0.45, max(0.2, n_rows * 0.025))
            gs = fig.add_gridspec(2, 1, height_ratios=[table_height_ratio, 1 - table_height_ratio], hspace=0.3)
            ax_table = fig.add_subplot(gs[0, 0])
            ax_table.axis("off")
            ax_table.axis("tight")

            if config.title:
                fig.suptitle(config.title, fontsize=styles.get("title_font_size", 18), fontweight="bold", y=0.98)

            # 如果分页了，在表格上方加提示
            if original_rows > MAX_ROWS:
                ax_table.text(0.5, 1.08, f"（共 {original_rows} 行，显示前 {MAX_ROWS} 行）",
                             transform=ax_table.transAxes, ha="center", fontsize=9, color="#666")

            ReportExporter._render_table_on_ax(ax_table, df, styles, font_size=base_font)

            n_charts = len(config.charts)
            cols = min(2, n_charts)
            rows = (n_charts + cols - 1) // cols
            gs_charts = gs[1, 0].subgridspec(rows, cols, hspace=0.35, wspace=0.25)
            for i, chart_cfg in enumerate(config.charts):
                ax_chart = fig.add_subplot(gs_charts[i // cols, i % cols])
                ReportExporter._render_chart_on_ax(ax_chart, chart_df, chart_cfg)

        elif has_charts:
            n_charts = len(config.charts)
            cols = min(2, n_charts)
            rows = (n_charts + cols - 1) // cols
            if config.title:
                fig.suptitle(config.title, fontsize=styles.get("title_font_size", 18), fontweight="bold", y=0.98)
            for i, chart_cfg in enumerate(config.charts):
                ax = fig.add_subplot(rows, cols, i + 1)
                ReportExporter._render_chart_on_ax(ax, chart_df, chart_cfg)

        else:
            ax = fig.add_subplot(111)
            ax.axis("off")
            ax.axis("tight")
            if config.title:
                ax.set_title(config.title, fontsize=styles.get("title_font_size", 18), fontweight="bold", pad=20)
            # 如果分页了，在标题下方加提示
            if original_rows > MAX_ROWS:
                ax.text(0.5, 1.02, f"共 {original_rows} 行，显示前 {MAX_ROWS} 行",
                       transform=ax.transAxes, ha="center", fontsize=10, color="#666")
            ReportExporter._render_table_on_ax(ax, df, styles, font_size=base_font)

        plt.tight_layout(rect=[0, 0, 1, 0.96] if config.title else None)

        buf = io.BytesIO()
        fig.savefig(buf, format="png", bbox_inches="tight", dpi=DPI)
        buf.seek(0)
        plt.close(fig)
        return buf.getvalue()

    @staticmethod
    def _render_table_on_ax(ax, df: pd.DataFrame, styles: Dict, font_size: Optional[float] = None):
        """在指定 Axes 上渲染表格"""
        n_rows, n_cols = df.shape
        table = ax.table(
            cellText=df.values,
            colLabels=df.columns if n_rows > 0 else None,
            cellLoc="center",
            loc="center",
        )
        table.auto_set_font_size(False)
        table.set_fontsize(font_size if font_size is not None else styles.get("font_size", 11))
        table.scale(1, 2.0)

        header_bg = styles.get("header_bg", "#2E86AB")
        header_color = styles.get("header_color", "#FFFFFF")
        grid_color = styles.get("grid_color", "#E0E4E8")

        for key, cell in table.get_celld().items():
            row, col = key
            cell.set_edgecolor(grid_color)
            cell.set_linewidth(0.5)
            if row == 0:
                cell.set_facecolor(header_bg)
                cell.set_text_props(color=header_color, fontweight="bold")
            else:
                cell.set_facecolor("#FAFBFC" if row % 2 == 0 else "#FFFFFF")

    @staticmethod
    def _render_chart_on_ax(ax, df: pd.DataFrame, chart_cfg: ChartConfig):
        """在指定 Axes 上渲染单个图表（简化版，复用 ChartBuilder 逻辑）"""
        if df.empty or not chart_cfg.x_field or not chart_cfg.y_field:
            ax.text(0.5, 0.5, "图表字段配置错误", ha="center", va="center", fontsize=10)
            ax.axis("off")
            return

        if chart_cfg.x_field not in df.columns or (chart_cfg.y_field != "__count__" and chart_cfg.y_field not in df.columns):
            ax.text(0.5, 0.5, "图表字段不存在", ha="center", va="center", fontsize=10)
            ax.axis("off")
            return
        colors = ChartBuilder.COLOR_SCHEMES.get(chart_cfg.color_scheme, ChartBuilder.COLOR_SCHEMES["default"])
        render_df = ChartBuilder.prepare_frame(df, chart_cfg)
        pivot_data = ChartBuilder.aggregate_frame(render_df, chart_cfg)

        # top_n
        if chart_cfg.top_n > 0 and len(pivot_data) > chart_cfg.top_n:
            if isinstance(pivot_data, pd.DataFrame):
                top_idx = pivot_data.sum(axis=1).nlargest(chart_cfg.top_n).index
            else:
                top_idx = pivot_data.nlargest(chart_cfg.top_n).index
            pivot_data = pivot_data.loc[top_idx]

        if chart_cfg.chart_type == "bar":
            if isinstance(pivot_data, pd.DataFrame) and pivot_data.shape[1] > 1:
                pivot_data.plot(kind="bar", ax=ax, color=colors, width=0.8)
            else:
                vals = pivot_data.values.flatten() if isinstance(pivot_data, pd.DataFrame) else pivot_data.values
                bars = ax.bar(range(len(pivot_data)), vals, color=colors[0], width=0.6)
                ax.set_xticks(range(len(pivot_data)))
                ax.set_xticklabels(pivot_data.index, rotation=45, ha="right")
                if chart_cfg.show_values:
                    for bar in bars:
                        h = bar.get_height()
                        ax.text(bar.get_x() + bar.get_width()/2., h, f"{h:.0f}", ha="center", va="bottom", fontsize=8)

        elif chart_cfg.chart_type == "horizontal_bar":
            if isinstance(pivot_data, pd.DataFrame) and pivot_data.shape[1] > 1:
                pivot_data.plot(kind="barh", ax=ax, color=colors, width=0.8)
            else:
                vals = pivot_data.values.flatten() if isinstance(pivot_data, pd.DataFrame) else pivot_data.values
                bars = ax.barh(range(len(pivot_data)), vals, color=colors[0], height=0.6)
                ax.set_yticks(range(len(pivot_data)))
                ax.set_yticklabels(pivot_data.index)
                if chart_cfg.show_values:
                    for bar in bars:
                        w = bar.get_width()
                        ax.text(w, bar.get_y() + bar.get_height()/2., f"{w:.0f}", ha="left", va="center", fontsize=8)

        elif chart_cfg.chart_type in ("pie", "donut"):
            vals = pivot_data.values.flatten() if isinstance(pivot_data, pd.DataFrame) else pivot_data.values
            wedgeprops = dict(width=0.4) if chart_cfg.chart_type == "donut" else None
            wedges, texts, autotexts = ax.pie(vals, labels=pivot_data.index, autopct="%1.1f%%",
                                               colors=colors, startangle=90, wedgeprops=wedgeprops)
            for autotext in autotexts:
                autotext.set_fontsize(8)

        elif chart_cfg.chart_type == "line":
            if isinstance(pivot_data, pd.DataFrame) and pivot_data.shape[1] > 1:
                for i, col in enumerate(pivot_data.columns):
                    ax.plot(pivot_data.index, pivot_data[col], marker="o", label=col, color=colors[i % len(colors)])
                ax.legend(fontsize=8)
            else:
                vals = pivot_data.values.flatten() if isinstance(pivot_data, pd.DataFrame) else pivot_data.values
                ax.plot(pivot_data.index, vals, marker="o", color=colors[0], linewidth=2)
            ax.set_xticklabels(pivot_data.index, rotation=45, ha="right")

        elif chart_cfg.chart_type == "area":
            if isinstance(pivot_data, pd.DataFrame) and pivot_data.shape[1] > 1:
                pivot_data.plot(kind="area", ax=ax, color=colors, alpha=0.6)
            else:
                vals = pivot_data.values.flatten() if isinstance(pivot_data, pd.DataFrame) else pivot_data.values
                ax.fill_between(range(len(pivot_data)), vals, alpha=0.6, color=colors[0])
                ax.plot(range(len(pivot_data)), vals, color=colors[0], linewidth=2)
                ax.set_xticks(range(len(pivot_data)))
                ax.set_xticklabels(pivot_data.index, rotation=45, ha="right")

        elif chart_cfg.chart_type == "scatter":
            if chart_cfg.group_field and chart_cfg.group_field in render_df.columns:
                for index, (name, subset) in enumerate(render_df.groupby(chart_cfg.group_field, observed=True, dropna=False)):
                    ax.scatter(subset[chart_cfg.x_field], subset[chart_cfg.y_field], alpha=0.6, color=colors[index % len(colors)], s=30, label=str(name))
                ax.legend(fontsize=8)
            else:
                ax.scatter(render_df[chart_cfg.x_field], render_df[chart_cfg.y_field], alpha=0.6, color=colors[0], s=30)
            ax.set_xlabel(chart_cfg.x_field, fontsize=9)

        ax.set_title(chart_cfg.title, fontsize=11, fontweight="bold", pad=8)
        y_label = "记录数" if chart_cfg.y_field == "__count__" else chart_cfg.y_field
        ax.set_ylabel(y_label if chart_cfg.chart_type not in ("pie", "donut") else "", fontsize=9)
        ax.set_xlabel(chart_cfg.x_field if chart_cfg.chart_type not in ("pie", "donut") else "", fontsize=9)
        ax.grid(axis="y", alpha=0.3, linestyle="--")
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.tick_params(labelsize=8)

    @staticmethod
    def to_excel(df: pd.DataFrame, config: ReportConfig, path: str):
        """导出为 Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError("请安装 openpyxl: pip install openpyxl")

        styles = config.styles
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="报表", index=False, startrow=1)
            ws = writer.sheets["报表"]

            # 标题
            if config.title:
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
                title_cell = ws.cell(row=1, column=1)
                title_cell.value = config.title
                title_cell.font = Font(size=16, bold=True)
                title_cell.alignment = Alignment(horizontal="center", vertical="center")

            def _clean_color(c):
                c = (c or "").strip()
                if not c:
                    return "000000"
                c = c.lstrip("#")
                if len(c) == 3:
                    c = "".join([ch*2 for ch in c])
                if len(c) == 6:
                    return c
                if len(c) == 8:
                    return c[2:]  # aRGB -> RGB
                return "000000"

            # 表头样式
            header_bg = _clean_color(styles.get("header_bg", "#2E86AB"))
            header_color = _clean_color(styles.get("header_color", "#FFFFFF"))
            grid_color = _clean_color(styles.get("grid_color", "#E0E4E8"))

            header_fill = PatternFill(
                start_color=header_bg,
                end_color=header_bg,
                fill_type="solid",
            )
            header_font = Font(color=header_color, bold=True)
            thin_border = Border(
                left=Side(style="thin", color=grid_color),
                right=Side(style="thin", color=grid_color),
                top=Side(style="thin", color=grid_color),
                bottom=Side(style="thin", color=grid_color),
            )

            header_row = 2 if config.title else 1
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            # 数据行样式
            for row in range(header_row + 1, header_row + len(df) + 1):
                for col in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if (row - header_row) % 2 == 0:
                        cell.fill = PatternFill(start_color="FAFBFC", end_color="FAFBFC", fill_type="solid")

            # 自动调整列宽
            for col in range(1, len(df.columns) + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                max_length = 0
                for row in range(1, header_row + len(df) + 1):
                    cell_value = str(ws.cell(row=row, column=col).value or "")
                    max_length = max(max_length, len(cell_value))
                ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

    @staticmethod
    def to_pdf(df: pd.DataFrame, config: ReportConfig, path: str):
        """导出为 PDF"""
        n_rows, n_cols = df.shape
        if n_rows == 0:
            n_rows, n_cols = 1, 1

        cell_height = 0.4
        cell_width = 1.8
        title_height = 0.6 if config.title else 0
        fig_height = max(4, n_rows * cell_height + title_height + 1)
        fig_width = max(8, n_cols * cell_width + 1)

        with PdfPages(path) as pdf:
            fig, ax = plt.subplots(figsize=(fig_width, fig_height))
            ax.axis("off")
            ax.axis("tight")

            if config.title:
                ax.set_title(
                    config.title,
                    fontsize=config.styles.get("title_font_size", 16),
                    fontweight="bold",
                    pad=20,
                )

            table = ax.table(
                cellText=df.values,
                colLabels=df.columns if n_rows > 0 else None,
                cellLoc="center",
                loc="center",
            )
            table.auto_set_font_size(False)
            table.set_fontsize(config.styles.get("font_size", 11))
            table.scale(1, 2)

            header_bg = config.styles.get("header_bg", "#2E86AB")
            header_color = config.styles.get("header_color", "#FFFFFF")
            grid_color = config.styles.get("grid_color", "#E0E4E8")

            for key, cell in table.get_celld().items():
                row, col = key
                cell.set_edgecolor(grid_color)
                cell.set_linewidth(0.5)
                if row == 0:
                    cell.set_facecolor(header_bg)
                    cell.set_text_props(color=header_color, fontweight="bold")
                else:
                    cell.set_facecolor("#FAFBFC" if row % 2 == 0 else "#FFFFFF")

            plt.tight_layout()
            pdf.savefig(fig, bbox_inches="tight", dpi=150)
            plt.close(fig)


class ReportEngine:
    """报表引擎入口"""

    @staticmethod
    def render(df: pd.DataFrame, config: ReportConfig) -> pd.DataFrame:
        """渲染报表，返回用于展示/导出的 DataFrame"""
        if config.mode == "pivot":
            return PivotBuilder.build(df, config.pivot)
        else:
            return FreeReportBuilder.build(df, config.cells)

    @staticmethod
    def preview(df: pd.DataFrame, config: ReportConfig) -> bytes:
        """渲染为图片字节（表格 + 图表组合）"""
        result_df = ReportEngine.render(df, config)
        # 图表需要原始数据，透视结果用于表格
        return ReportExporter.to_image(result_df, config, chart_df=df)

    @staticmethod
    def export(df: pd.DataFrame, config: ReportConfig, format: str, path: str):
        """导出报表"""
        result_df = ReportEngine.render(df, config)
        if format == "excel":
            ReportExporter.to_excel(result_df, config, path)
        elif format == "pdf":
            # PDF 也支持图表组合
            ReportExporter.to_pdf(result_df, config, path)
        else:
            raise ValueError(f"不支持的导出格式: {format}")
